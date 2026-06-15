#!/usr/bin/env python3
r"""
Inspect the four CRI bronze intake workbooks and print a JSON summary.

Run from project root with the CRI data_system venv Python, for example:
    .\ψ\incubate\DCCE\CRI\data_system\.venv\Scripts\python.exe .\ψ\incubate\DCCE\CRI\data_system\script\inspect_cri_workbooks.py
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


SCRIPT_PATH = Path(__file__).resolve()
PROJECT_ROOT = SCRIPT_PATH.parents[6]
BRONZE_DIR = PROJECT_ROOT / "ψ" / "incubate" / "DCCE" / "CRI" / "data_system" / "data" / "0_bronze" / "2026-06-12_cri_proj_data"

TARGET_WORKBOOKS = [
    BRONZE_DIR / "CRI Data - Eco loss.xlsx",
    BRONZE_DIR / "CRI Data - GPP.xlsx",
    BRONZE_DIR / "CRI Data - Heatwave.xlsx",
    BRONZE_DIR / "CRI Data - Population.xlsx",
]

DICTIONARY_SHEET_KEYWORDS = (
    "dict",
    "dictionary",
    "data dict",
    "metadata",
    "lookup",
    "codebook",
    "legend",
    "description",
    "variable",
    "indicator",
)

HEADER_HINT_KEYWORDS = {
    "province",
    "district",
    "amphoe",
    "tambon",
    "year",
    "month",
    "date",
    "code",
    "id",
    "name",
    "value",
    "unit",
    "indicator",
    "population",
    "gpp",
    "loss",
    "heatwave",
}


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def row_cells(ws, row_number: int, max_cols: int = 20) -> list[str]:
    values: list[str] = []
    for row in ws.iter_rows(min_row=row_number, max_row=row_number, min_col=1, max_col=max_cols, values_only=True):
        for cell in row:
            values.append(normalize_text(cell))
    return values


def score_header_candidate(values: list[str]) -> int:
    non_empty = [value for value in values if value]
    if not non_empty:
        return -1

    score = 0
    if len(non_empty) >= 2:
        score += 2
    if len(non_empty) >= 4:
        score += 2

    unique_ratio = len(set(non_empty)) / max(len(non_empty), 1)
    if unique_ratio > 0.8:
        score += 2

    joined = " ".join(non_empty).casefold()
    for keyword in HEADER_HINT_KEYWORDS:
        if keyword in joined:
            score += 2

    shortish = sum(1 for value in non_empty if len(value) <= 40)
    if shortish == len(non_empty):
        score += 2

    return score


def probable_dictionary_sheet(sheet_name: str, first_rows: list[list[str]]) -> bool:
    lowered = sheet_name.casefold()
    if any(keyword in lowered for keyword in DICTIONARY_SHEET_KEYWORDS):
        return True

    joined = " ".join(" ".join(row).casefold() for row in first_rows)
    if "description" in joined and "variable" in joined:
        return True
    if "indicator" in joined and "unit" in joined:
        return True
    return False


def inspect_workbook(path: Path) -> dict:
    result: dict = {
        "workbook_path": str(path),
        "exists": path.exists(),
        "sheets": [],
    }

    if not path.exists():
        result["error"] = "Workbook not found"
        return result

    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        result["sheet_count"] = len(workbook.sheetnames)
        result["sheet_names"] = list(workbook.sheetnames)

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            preview_rows = [row_cells(ws, row_number) for row_number in range(1, min(ws.max_row, 10) + 1)]

            header_candidates = []
            for row_number in range(1, min(ws.max_row, 25) + 1):
                values = row_cells(ws, row_number)
                score = score_header_candidate(values)
                if score >= 0:
                    header_candidates.append(
                        {
                            "row_number": row_number,
                            "score": score,
                            "values": values[:12],
                        }
                    )

            header_candidates.sort(key=lambda item: (-item["score"], item["row_number"]))

            result["sheets"].append(
                {
                    "sheet_name": sheet_name,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "is_probable_dictionary_sheet": probable_dictionary_sheet(sheet_name, preview_rows),
                    "top_preview_rows": preview_rows[:5],
                    "header_candidates": header_candidates[:5],
                }
            )
    finally:
        workbook.close()

    return result


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    report = {
        "bronze_dir": str(BRONZE_DIR),
        "workbooks": [inspect_workbook(path) for path in TARGET_WORKBOOKS],
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
