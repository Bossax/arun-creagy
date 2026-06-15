#!/usr/bin/env python3
r"""
Extract Population data tables from the CRI bronze intake workbook.

Run from project root with the CRI data_system venv Python, for example:
    .\ψ\incubate\DCCE\CRI\data_system\.venv\Scripts\python.exe .\ψ\incubate\DCCE\CRI\data_system\script\extract_population_tables.py

Outputs:
- `population_extracts/pop67.raw.csv`
- `population_extracts/pop60-67.raw.csv`
- `population_extracts/population.manifest.json`
"""

from __future__ import annotations

import csv
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


SCRIPT_PATH = Path(__file__).resolve()
PROJECT_ROOT = SCRIPT_PATH.parents[6]
BRONZE_DIR = PROJECT_ROOT / "ψ" / "incubate" / "DCCE" / "CRI" / "data_system" / "data" / "0_bronze" / "2026-06-12_cri_proj_data"
WORKBOOK_PATH = BRONZE_DIR / "CRI Data - Population.xlsx"
OUTPUT_DIR = BRONZE_DIR / "population_extracts"
MANIFEST_PATH = OUTPUT_DIR / "population.manifest.json"

SHEET_OUTPUTS = {
    "pop67": "pop67.raw.csv",
    "pop60-67": "pop60-67.raw.csv",
}

YEAR_PATTERN = re.compile(r"25\d{2}")
HEADER_TEXT_MARKERS = (
    "จำนวนประชากร",
    "รหัสจังหวัด",
    "ชื่อจังหวัด",
    "รหัสสำนัก",
    "รหัสสำนักทะเบียน",
    "ชื่อสำนักทะเบียน",
    "รหัสตำบล",
    "ชื่อตำบล",
)


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def worksheet_rows(ws) -> list[tuple[int, list[str]]]:
    rows: list[tuple[int, list[str]]] = []
    for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column, values_only=True), start=1):
        rows.append((row_number, [normalize_text(cell) for cell in row]))
    return rows


def count_non_empty(values: list[str]) -> int:
    return sum(1 for value in values if value)


def trim_trailing_empty_columns(rows: list[list[str]]) -> list[list[str]]:
    if not rows:
        return rows

    last_non_empty_index = -1
    for row in rows:
        for index, value in enumerate(row):
            if value:
                last_non_empty_index = max(last_non_empty_index, index)

    if last_non_empty_index < 0:
        return [[] for _ in rows]

    width = last_non_empty_index + 1
    return [row[:width] for row in rows]


def clean_headers(values: list[str]) -> list[str]:
    headers: list[str] = []
    used: dict[str, int] = {}

    for index, value in enumerate(values, start=1):
        header = normalize_text(value) or f"column_{index}"
        count = used.get(header, 0)
        used[header] = count + 1
        if count:
            header = f"{header}_{count + 1}"
        headers.append(header)

    return headers


def merge_header_rows(header_rows: list[list[str]]) -> list[str]:
    width = max((len(row) for row in header_rows), default=0)
    merged: list[str] = []

    for col_index in range(width):
        parts: list[str] = []
        seen: set[str] = set()
        for row in header_rows:
            value = normalize_text(row[col_index] if col_index < len(row) else "")
            if not value or value in seen:
                continue
            seen.add(value)
            parts.append(value)
        if any(is_year(part) for part in parts):
            year_parts = [part for part in parts if is_year(part)]
            non_year_parts = [part for part in parts if not is_year(part)]
            if len(year_parts) == 1 and len(non_year_parts) == 1 and "จำนวนประชากร" in non_year_parts[0]:
                merged.append(year_parts[0])
                continue
        merged.append(" | ".join(parts))

    return merged


def is_year(value: str) -> bool:
    return bool(YEAR_PATTERN.fullmatch(normalize_text(value)))


def count_years(row: list[str]) -> int:
    return sum(1 for value in row if is_year(value))


def has_header_text(row: list[str]) -> bool:
    joined = " ".join(row)
    return any(marker in joined for marker in HEADER_TEXT_MARKERS)


def is_footer_or_note_row(row: list[str]) -> bool:
    non_empty = [value for value in row if value]
    if not non_empty:
        return False

    joined = " ".join(non_empty).casefold()
    footer_markers = (
        "source",
        "note",
        "หมายเหตุ",
        "ที่มา",
        "unit",
        "หน่วย",
        "สำนักบริหารการทะเบียน",
        "กรมการปกครอง",
        "กระทรวงมหาดไทย",
    )
    return any(marker in joined for marker in footer_markers)


def detect_header_index(indexed_rows: list[tuple[int, list[str]]]) -> int:
    for index, (_, row) in enumerate(indexed_rows):
        if count_years(row) >= 4:
            previous_row = indexed_rows[index - 1][1] if index > 0 else []
            if previous_row and has_header_text(previous_row):
                return index - 1
            return index

    for index, (_, row) in enumerate(indexed_rows):
        if has_header_text(row):
            return index

    return 0


def detect_header_depth(indexed_rows: list[tuple[int, list[str]]], header_index: int) -> int:
    header_row = indexed_rows[header_index][1]
    next_row = indexed_rows[header_index + 1][1] if header_index + 1 < len(indexed_rows) else []

    if count_years(header_row) >= 4 and count_non_empty(next_row) >= 4:
        return 2

    if "จำนวนประชากร" in " ".join(header_row) and count_years(next_row) >= 4:
        return 2

    return 1


def looks_like_data_row(row: list[str]) -> bool:
    if not row:
        return False
    if count_non_empty(row) < 3:
        return False

    first = normalize_text(row[0]) if len(row) >= 1 else ""
    second = normalize_text(row[1]) if len(row) >= 2 else ""
    if re.fullmatch(r"\d{1,2}", first) and second:
        return True
    if re.fullmatch(r"\d{1,2}", first) and len(row) >= 3 and normalize_text(row[2]):
        return True
    return count_years(row) >= 1 and any(re.fullmatch(r"\d+", value) or value.startswith("=") for value in row)


def first_data_index(indexed_rows: list[tuple[int, list[str]]], start_index: int) -> int:
    for index in range(start_index, len(indexed_rows)):
        row = indexed_rows[index][1]
        if is_footer_or_note_row(row):
            continue
        if looks_like_data_row(row):
            return index
    return start_index


def collect_formula_samples(data_rows: list[list[str]], limit: int = 10) -> list[dict[str, object]]:
    samples: list[dict[str, object]] = []
    for row_offset, row in enumerate(data_rows, start=1):
        for col_offset, value in enumerate(row, start=1):
            if value.startswith("="):
                samples.append(
                    {
                        "data_row_index_1_based": row_offset,
                        "column_index_1_based": col_offset,
                        "formula": value,
                    }
                )
                if len(samples) >= limit:
                    return samples
    return samples


def structural_issues(headers: list[str], data_rows: list[list[str]]) -> list[str]:
    issues: list[str] = []
    if not headers:
        issues.append("No headers detected")
    if not data_rows:
        issues.append("No data rows extracted")
        return issues

    expected_width = len(headers)
    inconsistent_rows = sum(1 for row in data_rows if len(row) != expected_width)
    if inconsistent_rows:
        issues.append(f"{inconsistent_rows} data rows have width different from header width {expected_width}")

    if any("column_" in header for header in headers):
        issues.append("One or more header cells were blank and replaced with synthetic column names")

    return issues


def extract_sheet(ws, csv_name: str) -> dict:
    indexed_rows = worksheet_rows(ws)
    indexed_rows = [(row_number, row) for row_number, row in indexed_rows if any(row)]
    indexed_rows = [
        (row_number, row)
        for row_number, row in zip(
            [row_number for row_number, _ in indexed_rows],
            trim_trailing_empty_columns([row for _, row in indexed_rows]),
        )
    ]

    if not indexed_rows:
        return {
            "sheet_name": ws.title,
            "error": "No non-empty rows found",
            "header_row_number": 1,
        }

    header_index = detect_header_index(indexed_rows)
    header_depth = detect_header_depth(indexed_rows, header_index)
    header_entries = indexed_rows[header_index:header_index + header_depth]
    header_rows = [row for _, row in header_entries]
    headers = clean_headers(merge_header_rows(header_rows))

    data_start_index = first_data_index(indexed_rows, header_index + header_depth)
    data_rows: list[list[str]] = []
    footer_rows: list[list[str]] = []

    for _, row in indexed_rows[data_start_index:]:
        if is_footer_or_note_row(row):
            footer_rows.append(row)
            continue
        data_rows.append(row)

    while data_rows and not any(data_rows[-1]):
        data_rows.pop()

    csv_path = OUTPUT_DIR / csv_name
    with csv_path.open("w", encoding="utf-8-sig", newline="") as fp:
        writer = csv.writer(fp)
        writer.writerow(headers)
        writer.writerows(data_rows)

    formula_samples = collect_formula_samples(data_rows)
    return {
        "sheet_name": ws.title,
        "output_csv_path": str(csv_path),
        "worksheet_dimensions": {
            "max_row": ws.max_row,
            "max_column": ws.max_column,
        },
        "header_row_number": header_entries[0][0],
        "header_row_span": header_depth,
        "header_rows": header_rows,
        "headers": headers,
        "column_count": len(headers),
        "row_count": len(data_rows),
        "preview_first_data_row": data_rows[0] if data_rows else [],
        "preview_second_data_row": data_rows[1] if len(data_rows) > 1 else [],
        "footer_rows_skipped": footer_rows,
        "formula_cells_detected": sum(1 for row in data_rows for value in row if value.startswith("=")),
        "formula_samples": formula_samples,
        "raw_formula_preservation_policy": "Bronze preserves raw formulas exactly as workbook cell text; resolved values belong in Silver.",
        "structural_issues": structural_issues(headers, data_rows),
    }


def extract_tables() -> dict:
    manifest: dict = {
        "script_path": str(SCRIPT_PATH),
        "workbook_path": str(WORKBOOK_PATH),
        "output_manifest_path": str(MANIFEST_PATH),
        "output_dir": str(OUTPUT_DIR),
        "extracted_at_utc": datetime.now(timezone.utc).isoformat(),
        "exists": WORKBOOK_PATH.exists(),
        "bronze_policy": "Preserve raw formulas in Bronze and derive resolved values later in Silver.",
        "sheets": [],
    }

    if not WORKBOOK_PATH.exists():
        manifest["error"] = "Workbook not found"
        return manifest

    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=False)
    try:
        manifest["sheet_names"] = list(workbook.sheetnames)
        for sheet_name, csv_name in SHEET_OUTPUTS.items():
            if sheet_name not in workbook.sheetnames:
                manifest["sheets"].append(
                    {
                        "sheet_name": sheet_name,
                        "output_csv_path": str(OUTPUT_DIR / csv_name),
                        "error": "Sheet not found",
                    }
                )
                continue
            manifest["sheets"].append(extract_sheet(workbook[sheet_name], csv_name))
        return manifest
    finally:
        workbook.close()


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    manifest = extract_tables()
    MANIFEST_PATH.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
