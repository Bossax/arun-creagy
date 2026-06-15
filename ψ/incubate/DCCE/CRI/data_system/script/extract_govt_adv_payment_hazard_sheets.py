#!/usr/bin/env python3
r"""
Extract Government Advance Payment hazard-sheet tables from the CRI bronze intake workbook.

Run from project root with the CRI data_system venv Python, for example:
    .\ψ\incubate\DCCE\CRI\data_system\.venv\Scripts\python.exe .\ψ\incubate\DCCE\CRI\data_system\script\extract_govt_adv_payment_hazard_sheets.py

Outputs:
- `govt_adv_payment_extracts/govt_adv_payment-อุทกภัย.raw.csv`
- `govt_adv_payment_extracts/govt_adv_payment-ภัยแล้ง.raw.csv`
- `govt_adv_payment_extracts/govt_adv_payment-วาตภัย.raw.csv`
- `govt_adv_payment_extracts/govt_adv_payment.manifest.json`
"""

from __future__ import annotations

import csv
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


SCRIPT_PATH = Path(__file__).resolve()
PROJECT_ROOT = SCRIPT_PATH.parents[6]
BRONZE_DIR = PROJECT_ROOT / "ψ" / "incubate" / "DCCE" / "CRI" / "data_system" / "data" / "0_bronze" / "2026-06-12_cri_proj_data"
WORKBOOK_PATH = BRONZE_DIR / "CRI Data - Government_Advanced_Payment.xlsx"
OUTPUT_DIR = BRONZE_DIR / "govt_adv_payment_extracts"
MANIFEST_PATH = OUTPUT_DIR / "govt_adv_payment.manifest.json"

SHEET_OUTPUTS = {
    "Eco loss อุทกภัย": "govt_adv_payment-อุทกภัย.raw.csv",
    "Eco loss ภัยแล้ง": "govt_adv_payment-ภัยแล้ง.raw.csv",
    "Eco loss วาตภัย": "govt_adv_payment-วาตภัย.raw.csv",
}


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def row_values(ws, row_number: int, max_cols: int | None = None) -> list[str]:
    upper = max_cols or ws.max_column
    for row in ws.iter_rows(min_row=row_number, max_row=row_number, min_col=1, max_col=upper, values_only=True):
        return [normalize_text(cell) for cell in row]
    return []


def count_non_empty(values: list[str]) -> int:
    return sum(1 for value in values if value)


def trim_trailing_empty_columns(rows: list[list[str]]) -> list[list[str]]:
    if not rows:
        return rows

    last_non_empty_index = -1
    for row in rows:
        for index, value in enumerate(row):
            if value:
                last_non_empty_index = max(last_non_empty_index, index)

    if last_non_empty_index < 0:
        return [[] for _ in rows]

    width = last_non_empty_index + 1
    return [row[:width] for row in rows]


def clean_headers(values: list[str]) -> list[str]:
    headers: list[str] = []
    used: dict[str, int] = {}

    for index, value in enumerate(values, start=1):
        header = normalize_text(value) or f"column_{index}"
        count = used.get(header, 0)
        used[header] = count + 1
        if count:
            header = f"{header}_{count + 1}"
        headers.append(header)

    return headers


def merge_header_rows(header_rows: list[list[str]]) -> list[str]:
    width = max((len(row) for row in header_rows), default=0)
    merged: list[str] = []

    for col_index in range(width):
        parts: list[str] = []
        seen: set[str] = set()
        for row in header_rows:
            value = normalize_text(row[col_index] if col_index < len(row) else "")
            if not value or value in seen:
                continue
            seen.add(value)
            parts.append(value)
        merged.append(" | ".join(parts))

    return merged


def is_numeric_like(value: str) -> bool:
    return bool(re.fullmatch(r"[-+]?\d+(?:\.\d+)?(?:\s*-\s*\d+(?:\.\d+)?)?", value))


def is_hazard_sheet_title(value: str) -> bool:
    normalized = normalize_text(value).casefold()
    return normalized.startswith("eco loss") or normalized.startswith("govt_adv_payment")


def is_data_row_leader(value: str) -> bool:
    normalized = normalize_text(value)
    if not normalized:
        return False
    if re.fullmatch(r"\d{2}", normalized):
        return True
    if re.fullmatch(r"\d{2}\s*-\s*.+", normalized):
        return True
    if re.fullmatch(r"\d+", normalized):
        return True
    return False


def row_has_years(row: list[str]) -> bool:
    return sum(1 for value in row if re.fullmatch(r"(?:19|20|25)\d{2}", normalize_text(value))) >= 2


def row_has_thai_labels(row: list[str]) -> bool:
    return sum(1 for value in row if value and re.search(r"[ก-๙A-Za-z]", value)) >= 2


def row_is_footer_or_note(row: list[str]) -> bool:
    non_empty = [value for value in row if value]
    if not non_empty:
        return False

    joined = " ".join(non_empty).casefold()
    footer_markers = (
        "source",
        "note",
        "หมายเหตุ",
        "ที่มา",
        "หน่วย",
        "unit",
        "ล้านบาท",
        "กรม",
        "สำนักงาน",
    )
    if any(marker in joined for marker in footer_markers):
        return True

    if len(non_empty) <= 2 and any(marker in joined for marker in footer_markers):
        return True

    return False


def detect_header_start(indexed_rows: list[tuple[int, list[str]]]) -> int:
    for index, (_, row) in enumerate(indexed_rows):
        if not any(row):
            continue
        if row_has_years(row):
            return index
        if row_has_thai_labels(row) and any("จังหวัด" in value or "province" in value.casefold() for value in row if value):
            return index

    for index, (_, row) in enumerate(indexed_rows):
        if any(row) and not is_hazard_sheet_title(next((value for value in row if value), "")):
            return index
    return 0


def detect_header_depth(rows: list[list[str]]) -> int:
    if len(rows) < 2:
        return 1

    first_row = rows[0]
    second_row = rows[1]
    third_row = rows[2] if len(rows) >= 3 else []

    if row_has_years(first_row) and row_has_thai_labels(second_row):
        return 2

    second_non_empty = count_non_empty(second_row)
    second_numeric = sum(1 for value in second_row if value and is_numeric_like(value))
    second_label_like = sum(1 for value in second_row if value and re.search(r"[A-Za-zก-๙]", value))
    third_non_empty = count_non_empty(third_row)
    third_label_like = sum(1 for value in third_row if value and re.search(r"[A-Za-zก-๙]", value))

    if third_non_empty >= 2 and third_label_like >= 2 and not any(is_data_row_leader(value) for value in third_row[:2]):
        return 3

    if second_non_empty >= 2 and second_label_like >= 2 and second_numeric < second_non_empty:
        return 2

    return 1


def detect_first_data_index(indexed_rows: list[tuple[int, list[str]]], header_index: int, header_depth: int) -> int:
    start = min(header_index + header_depth, len(indexed_rows))

    for index in range(start, len(indexed_rows)):
        _, row = indexed_rows[index]
        if row_is_footer_or_note(row):
            continue
        if is_data_row_leader(row[0] if row else ""):
            return index
        if len(row) >= 2 and is_data_row_leader(row[1]):
            return index
        if count_non_empty(row) >= 3 and not row_has_years(row):
            return index

    return start


def extract_sheet(ws, csv_name: str) -> dict:
    indexed_rows = [(row_number, row_values(ws, row_number)) for row_number in range(1, ws.max_row + 1)]
    indexed_rows = [(row_number, row) for row_number, row in indexed_rows if any(row)]
    indexed_rows = [
        (row_number, row)
        for row_number, row in zip(
            [row_number for row_number, _ in indexed_rows],
            trim_trailing_empty_columns([row for _, row in indexed_rows]),
        )
    ]

    if not indexed_rows:
        return {
            "sheet_name": ws.title,
            "error": "No non-empty rows found",
            "header_row_number": 1,
        }

    header_index = detect_header_start(indexed_rows)
    header_candidates = [row for _, row in indexed_rows[header_index:]]
    header_depth = detect_header_depth(header_candidates)
    header_rows = [row for _, row in indexed_rows[header_index:header_index + header_depth]]
    headers = clean_headers(merge_header_rows(header_rows))
    first_data_index = detect_first_data_index(indexed_rows, header_index, header_depth)

    data_rows: list[list[str]] = []
    footer_rows: list[list[str]] = []

    for _, row in indexed_rows[first_data_index:]:
        if row_is_footer_or_note(row):
            footer_rows.append(row)
            continue
        data_rows.append(row)

    while data_rows and not any(data_rows[-1]):
        data_rows.pop()

    csv_path = OUTPUT_DIR / csv_name
    with csv_path.open("w", encoding="utf-8-sig", newline="") as fp:
        writer = csv.writer(fp)
        writer.writerow(headers)
        writer.writerows(data_rows)

    return {
        "sheet_name": ws.title,
        "output_csv_path": str(csv_path),
        "worksheet_dimensions": {
            "max_row": ws.max_row,
            "max_column": ws.max_column,
        },
        "header_row_number": indexed_rows[header_index][0],
        "header_row_span": header_depth,
        "header_rows": header_rows,
        "headers": headers,
        "column_count": len(headers),
        "row_count": len(data_rows),
        "preview_first_data_row": data_rows[0] if data_rows else [],
        "footer_rows_skipped": footer_rows,
        "first_data_row_number": indexed_rows[first_data_index][0] if first_data_index < len(indexed_rows) else None,
    }


def extract_tables() -> dict:
    manifest: dict = {
        "workbook_path": str(WORKBOOK_PATH),
        "output_manifest_path": str(MANIFEST_PATH),
        "output_dir": str(OUTPUT_DIR),
        "exists": WORKBOOK_PATH.exists(),
        "sheets": [],
    }

    if not WORKBOOK_PATH.exists():
        manifest["error"] = "Workbook not found"
        return manifest

    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=False)
    try:
        manifest["sheet_names"] = list(workbook.sheetnames)

        for sheet_name, csv_name in SHEET_OUTPUTS.items():
            if sheet_name not in workbook.sheetnames:
                manifest["sheets"].append(
                    {
                        "sheet_name": sheet_name,
                        "output_csv_path": str(OUTPUT_DIR / csv_name),
                        "error": "Sheet not found",
                    }
                )
                continue

            ws = workbook[sheet_name]
            manifest["sheets"].append(extract_sheet(ws, csv_name))

        return manifest
    finally:
        workbook.close()


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    manifest = extract_tables()
    MANIFEST_PATH.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
