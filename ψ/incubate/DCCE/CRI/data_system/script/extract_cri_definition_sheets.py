#!/usr/bin/env python3
r"""
Extract the shared `Definition` sheets from the CRI bronze intake workbooks.

Run from project root with the CRI data_system venv Python, for example:
    .\ψ\incubate\DCCE\CRI\data_system\.venv\Scripts\python.exe .\ψ\incubate\DCCE\CRI\data_system\script\extract_cri_definition_sheets.py

Outputs:
- [`definition_sheet_extracts/manifest.json`](ψ/incubate/DCCE/CRI/data_system/data/0_bronze/2026-06-12_cri_proj_data/definition_sheet_extracts/manifest.json)
- one JSON file per workbook
- one Markdown review file per workbook
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


SCRIPT_PATH = Path(__file__).resolve()
PROJECT_ROOT = SCRIPT_PATH.parents[6]
BRONZE_DIR = PROJECT_ROOT / "ψ" / "incubate" / "DCCE" / "CRI" / "data_system" / "data" / "0_bronze" / "2026-06-12_cri_proj_data"
OUTPUT_DIR = BRONZE_DIR / "definition_sheet_extracts"

TARGET_WORKBOOKS = [
    BRONZE_DIR / "CRI Data - Government_Advanced_Payment.xlsx",
    BRONZE_DIR / "CRI Data - GPP.xlsx",
    BRONZE_DIR / "CRI Data - Heatwave.xlsx",
    BRONZE_DIR / "CRI Data - Population.xlsx",
]


def slugify(text: str) -> str:
    text = text.strip().casefold()
    text = text.replace("government_advanced_payment", "govt-adv-payment")
    text = re.sub(r"[^\w\s-]", "", text)
    text = re.sub(r"[-\s]+", "-", text)
    return text.strip("-") or "workbook"


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def extract_definition_rows(workbook_path: Path) -> dict:
    result: dict = {
        "workbook_path": str(workbook_path),
        "workbook_name": workbook_path.name,
        "definition_sheet_name": "Definition",
        "exists": workbook_path.exists(),
        "rows": [],
    }

    if not workbook_path.exists():
        result["error"] = "Workbook not found"
        return result

    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        if "Definition" not in workbook.sheetnames:
            result["error"] = "Definition sheet not found"
            result["sheet_names"] = list(workbook.sheetnames)
            return result

        ws = workbook["Definition"]
        result["max_row"] = ws.max_row
        result["max_column"] = ws.max_column

        for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
            cells = [normalize_text(cell) for cell in row]
            if not any(cells):
                continue

            term = cells[0] if len(cells) >= 1 else ""
            meaning = cells[1] if len(cells) >= 2 else ""

            result["rows"].append(
                {
                    "row_number": row_number,
                    "term": term,
                    "meaning": meaning,
                    "extra_cells": cells[2:],
                    "is_header": row_number == 1 and term.casefold() == "definition",
                }
            )
    finally:
        workbook.close()

    return result


def write_markdown_review(extract: dict, output_path: Path) -> None:
    lines = []
    lines.append(f"# Definition extract — {extract['workbook_name']}")
    lines.append("")
    lines.append(f"- Workbook: `{extract['workbook_path']}`")
    lines.append(f"- Sheet: `{extract['definition_sheet_name']}`")
    lines.append(f"- Exists: `{extract['exists']}`")
    if "error" in extract:
        lines.append(f"- Error: `{extract['error']}`")
        lines.append("")
        output_path.write_text("\n".join(lines), encoding="utf-8")
        return

    lines.append("")
    lines.append("| Row | Term | Meaning |")
    lines.append("|---|---|---|")
    for row in extract["rows"]:
        term = row["term"].replace("|", "\\|")
        meaning = row["meaning"].replace("|", "\\|")
        lines.append(f"| {row['row_number']} | {term} | {meaning} |")

    output_path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    manifest = {
        "source_root": str(BRONZE_DIR),
        "output_root": str(OUTPUT_DIR),
        "extracts": [],
    }

    for workbook_path in TARGET_WORKBOOKS:
        extract = extract_definition_rows(workbook_path)
        stem = slugify(workbook_path.stem)
        json_path = OUTPUT_DIR / f"{stem}.definition.json"
        md_path = OUTPUT_DIR / f"{stem}.definition.md"

        json_path.write_text(json.dumps(extract, ensure_ascii=False, indent=2), encoding="utf-8")
        write_markdown_review(extract, md_path)

        manifest["extracts"].append(
            {
                "workbook_name": workbook_path.name,
                "json_path": str(json_path),
                "markdown_path": str(md_path),
                "row_count": len(extract.get("rows", [])),
                "error": extract.get("error"),
            }
        )

    manifest_path = OUTPUT_DIR / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
