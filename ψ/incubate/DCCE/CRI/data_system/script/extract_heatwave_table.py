#!/usr/bin/env python3
r"""
Extract the Heatwave data table from the CRI bronze intake workbook.

Run from project root with the CRI data_system venv Python, for example:
    .\ψ\incubate\DCCE\CRI\data_system\.venv\Scripts\python.exe .\ψ\incubate\DCCE\CRI\data_system\script\extract_heatwave_table.py

Outputs:
- `heatwave_extracts/heatwave.raw.csv`
- `heatwave_extracts/heatwave.manifest.json`
"""

from __future__ import annotations

import csv
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


SCRIPT_PATH = Path(__file__).resolve()
PROJECT_ROOT = SCRIPT_PATH.parents[6]
BRONZE_DIR = PROJECT_ROOT / "ψ" / "incubate" / "DCCE" / "CRI" / "data_system" / "data" / "0_bronze" / "2026-06-12_cri_proj_data"
WORKBOOK_PATH = BRONZE_DIR / "CRI Data - Heatwave.xlsx"
SHEET_NAME = "Heatwave"
OUTPUT_DIR = BRONZE_DIR / "heatwave_extracts"
CSV_PATH = OUTPUT_DIR / "heatwave.raw.csv"
MANIFEST_PATH = OUTPUT_DIR / "heatwave.manifest.json"


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def row_values(ws, row_number: int, max_cols: int | None = None) -> list[str]:
    upper = max_cols or ws.max_column
    for row in ws.iter_rows(min_row=row_number, max_row=row_number, min_col=1, max_col=upper, values_only=True):
        return [normalize_text(cell) for cell in row]
    return []


def count_non_empty(values: list[str]) -> int:
    return sum(1 for value in values if value)


def find_header_row(ws) -> int:
    best_row = 1
    best_score = -1

    for row_number in range(1, min(ws.max_row, 40) + 1):
        values = row_values(ws, row_number)
        non_empty = [value for value in values if value]
        if not non_empty:
            continue

        score = 0
        score += count_non_empty(values) * 5

        unique_ratio = len(set(non_empty)) / len(non_empty)
        if unique_ratio > 0.8:
            score += 10

        alpha_like = sum(1 for value in non_empty if re.search(r"[A-Za-z]", value))
        numeric_like = sum(1 for value in non_empty if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", value))
        score += alpha_like * 2
        score -= numeric_like * 2

        joined = " ".join(non_empty).casefold()
        for keyword in ("province", "year", "month", "date", "heatwave", "severity", "days", "risk", "index", "name", "code"):
            if keyword in joined:
                score += 4

        if row_number <= 8:
            score += 8 - row_number

        if score > best_score:
            best_score = score
            best_row = row_number

    return best_row


def clean_headers(values: list[str]) -> list[str]:
    headers: list[str] = []
    used: dict[str, int] = {}

    for index, value in enumerate(values, start=1):
        header = normalize_text(value) or f"column_{index}"
        count = used.get(header, 0)
        used[header] = count + 1
        if count:
            header = f"{header}_{count + 1}"
        headers.append(header)

    return headers


def trim_trailing_empty_columns(rows: list[list[str]]) -> list[list[str]]:
    if not rows:
        return rows

    last_non_empty_index = -1
    for row in rows:
        for index, value in enumerate(row):
            if value:
                last_non_empty_index = max(last_non_empty_index, index)

    if last_non_empty_index < 0:
        return [[] for _ in rows]

    width = last_non_empty_index + 1
    return [row[:width] for row in rows]


def detect_header_depth(rows: list[list[str]]) -> int:
    if len(rows) < 2:
        return 1

    first_row = rows[0]
    second_row = rows[1]
    third_row = rows[2] if len(rows) >= 3 else []
    second_non_empty = count_non_empty(second_row)
    numeric_like = sum(1 for value in second_row if value and re.fullmatch(r"[-+]?\d+(?:\.\d+)?(?:\s*-\s*\d+(?:\.\d+)?)?", value))
    label_like = sum(1 for value in second_row if value and re.search(r"[A-Za-zก-๙]", value))
    third_non_empty = count_non_empty(third_row)
    repeated_metric_labels = sum(1 for value in third_row if value in {"Deaths", "Injured"})

    if third_non_empty >= 2 and repeated_metric_labels >= 2:
        return 3

    if second_non_empty >= 2 and label_like >= 2 and numeric_like < second_non_empty:
        return 2

    if any(value for value in first_row[2:]) and any(value for value in second_row[2:]):
        return 2

    return 1


def merge_header_rows(header_rows: list[list[str]]) -> list[str]:
    width = max((len(row) for row in header_rows), default=0)
    merged: list[str] = []

    for col_index in range(width):
        parts: list[str] = []
        seen: set[str] = set()
        for row in header_rows:
            value = normalize_text(row[col_index] if col_index < len(row) else "")
            if not value:
                continue
            if value in seen:
                continue
            seen.add(value)
            parts.append(value)
        merged.append(" | ".join(parts))

    return merged


def extract_table() -> dict:
    manifest: dict = {
        "workbook_path": str(WORKBOOK_PATH),
        "sheet_name": SHEET_NAME,
        "output_csv_path": str(CSV_PATH),
        "output_manifest_path": str(MANIFEST_PATH),
        "exists": WORKBOOK_PATH.exists(),
    }

    if not WORKBOOK_PATH.exists():
        manifest["error"] = "Workbook not found"
        return manifest

    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=False)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            manifest["error"] = "Heatwave sheet not found"
            manifest["sheet_names"] = list(workbook.sheetnames)
            return manifest

        ws = workbook[SHEET_NAME]
        manifest["worksheet_dimensions"] = {
            "max_row": ws.max_row,
            "max_column": ws.max_column,
        }

        header_row_number = find_header_row(ws)
        raw_rows = [row_values(ws, row_number) for row_number in range(header_row_number, ws.max_row + 1)]
        raw_rows = trim_trailing_empty_columns(raw_rows)
        non_empty_rows = [row for row in raw_rows if any(cell for cell in row)]

        if not non_empty_rows:
            manifest["error"] = "No non-empty rows found from detected header row"
            manifest["header_row_number"] = header_row_number
            return manifest

        header_depth = detect_header_depth(non_empty_rows)
        header_rows = non_empty_rows[:header_depth]
        headers = clean_headers(merge_header_rows(header_rows))
        data_rows = non_empty_rows[header_depth:]

        while data_rows and not any(data_rows[-1]):
            data_rows.pop()

        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        with CSV_PATH.open("w", encoding="utf-8-sig", newline="") as fp:
            writer = csv.writer(fp)
            writer.writerow(headers)
            writer.writerows(data_rows)

        manifest.update(
            {
                "header_row_number": header_row_number,
                "header_row_span": header_depth,
                "headers": headers,
                "column_count": len(headers),
                "row_count": len(data_rows),
                "preview_first_data_row": data_rows[0] if data_rows else [],
                "trailing_empty_rows_removed": True,
            }
        )
        return manifest
    finally:
        workbook.close()


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    manifest = extract_table()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    MANIFEST_PATH.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
